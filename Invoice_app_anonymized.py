import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
import tempfile
import os
import locale
from io import BytesIO
from xlsx2html import xlsx2html
import pdfkit
import yagmail
# Streamlit app
import numpy as np


def main():
    sender_email = '' #put your email address here
    sender_password = '' #put your password here (app_password if you have 2fa turned on)
    st.title('Invoice Generator')
    client_info = pd.read_excel('Info_clients.xlsx')
    # User input for Uurtarief
    #Uurtarief = st.session_state.Uurtarief = st.number_input('Enter Uurtarief', min_value=0.0, value=28.0)

    # User input for KM_vergoeding
    #Km_vergoeding = st.session_state.km_vergoeding = st.number_input('Enter KM vergoeding', min_value=0.0, value=0.19)

    # Upload Excel file
    uploaded_file = st.file_uploader('Upload your Excel file', type=['xlsx'])
    
    # Function to process the uploaded Excel file
    def process_excel_file(uploaded_file,client_info):
        df = pd.read_excel(uploaded_file, header=16)
        #st.dataframe(df)
        df['Naam Locatie'].replace('', np.nan, inplace=True)
        df.dropna(subset=['Naam Locatie'], inplace=True)
        df['Medewerker'][0] = df['Medewerker'][0].lower()
        client_info.iloc[:,1] = [x.lower() for x in client_info.iloc[:,1]]
        Uurtarief = st.session_state.Uurtarief = client_info.iloc[:,10][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        
        # User input for KM_vergoeding
        Km_vergoeding = st.session_state.km_vergoeding = client_info.iloc[:,11][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        mail = client_info.iloc[:,12][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
    # Variables
        Rekeningummer = client_info.iloc[:,8][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        KVK = client_info.iloc[:,7][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        Btw_nummer = client_info.iloc[:,9][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        Bedrijfsnaam = client_info.iloc[:,4][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]

       # if df["Medewerker"].str.contains("employer 1|employer 2").any():
       #     Rekeningummer = 'NL00 INGB 0000 0000 01'
       #     KVK = '123456'
       #  client_infoer = 'NL000000'
       #     Bedrijfsnaam = 'Company z'
       # else:
       #     Rekeningummer = 'NL00 INGB 0000 0000 02'
       #     KVK = '54321'
       #     Btw_nummer = 'NL000000002'
       #     Bedrijfsnaam = 'Company Q'


       # Create a dictionary to store user-provided kilometers per row
        km_input_dict = []
        default_km = 150
        distinct_locaties = df[df['Naam Locatie'] != 'Totaal']['Naam Locatie']

        # Create an input widget for each unique locatie
        for locatie in range(len(distinct_locaties)):
            km_input = st.number_input(f"{distinct_locaties[locatie]} - Noteer kilometers: {df['datum'][locatie]}", min_value=0, value=default_km, max_value=200, key=locatie)
            km_input_dict.append(km_input)
        #st.write(km_input_dict)
        # Update the DataFrame 'KM' column with the user-provided kilometers for corresponding 'Naam Locatie'
        #df['KM'] = df.apply(lambda row: km_input_dict[row['Naam Locatie']] if row['Naam Locatie'] in km_input_dict else row['KM'], axis=1)
        for x in range(len(df)-1):
            df.loc[x,'KM'] = km_input_dict[x]
        #df[df['KM'] != 'Totaal']['KM'] = km_input_dict


        # Add the user-provided kilometers to the DataFrame
        #df['KM'] = km_input
        df.dropna()
        df.drop(['Medewerker', 'Dag','Slaap of Waak?', 'Totaal'], axis=1)

        # Define the desired column order
        desired_order = ['datum', 'Naam Locatie', 1, 1.22, 1.38, 1.44, 1.49, 1.6, 'Begintijd', 'Eindtijd', 'uren maal toeslag','KM']

        # Reindex the dataframe with the desired column order
        df = df.reindex(columns=desired_order)

        #creating total hours columns
        df['uren maal toeslag'] = df[1]*1 + df[1.22]*1.22 + df[1.38]*1.38 + df[1.44]*1.44 + df[1.49]*1.49 + df[1.6]*1.6

        # Rounding 'uren totaal incl %' column to 2 decimal places
        df['uren maal toeslag'] = round(df['uren maal toeslag'], 2)

        # Formatting 'datum' column
        df['datum'] = pd.to_datetime(df['datum']).dt.strftime("%d-%m-%Y")
        # Formatting 'Begintijd' column
        df['Begintijd'] = pd.to_datetime(df['Begintijd'], format='%H:%M:%S').dt.strftime('%H:%M')
        # Formatting 'Eindtijd' column
        df['Eindtijd'] = pd.to_datetime(df['Eindtijd'], format='%H:%M:%S').dt.strftime('%H:%M')

    
        for i in range(len(df) - 1):
            current_row = df.iloc[i]
            previous_row = df.iloc[i - 1]
            
            if current_row['datum'] == previous_row['datum']:
                    df.at[i, 'KM'] = 0


        df = df.drop(df.index[-1])
    
    
        # Calculate the sum of 'uren maal toeslag' and 'KM' columns
        total_row = pd.DataFrame({
            'uren maal toeslag': df['uren maal toeslag'].sum(),
            'KM': df['KM'].sum(),
            'Eindtijd': 'Totaal'
        }, index=['Total'])

        # Append the total row to the input_file dataframe
        df = df.append(total_row)

        # Reset the index of the dataframe
        df.reset_index(drop=True, inplace=True)
        df = df.fillna('')

        # Set the locale to Dutch (Netherlands)
        locale.setlocale(locale.LC_TIME, 'nl_NL.UTF-8')

        # Get the current date
        current_date = datetime.now()

        # Extract year and month from the current date
        year = current_date.year
        month_number = current_date.month

        # Get the full month name based on the month number
        month_name = current_date.strftime('%B')

        # Generate the fac_number combining year and current month number
        fac_number = f"{year}{month_number:02d}"

        # Get the whole date today in your desired format
        date_today = current_date.strftime('%Y-%m-%d')
    
        ## Calculate the values for the table
        Totale_uren = total_row['uren maal toeslag'].iloc[0]
        Uren_maal_tarief = Totale_uren * Uurtarief
        Totale_km = total_row['KM'].iloc[0]
        KM_maal_tarief = Totale_km * Km_vergoeding
        sub_totaal = Uren_maal_tarief + KM_maal_tarief
        
        return mail, df, Totale_uren, Uurtarief, Uren_maal_tarief, Totale_km, Km_vergoeding, KM_maal_tarief, sub_totaal, Bedrijfsnaam, KVK, Rekeningummer, Btw_nummer, month_name, fac_number, date_today


    if uploaded_file is None:
        st.write("")

    else:
        name = pd.read_excel(uploaded_file, header=16)
        name = name['Medewerker'][0]
        
        # Process the uploaded file and get the variables
        mail, df, Totale_uren, Uurtarief, Uren_maal_tarief, Totale_km, Km_vergoeding, KM_maal_tarief, sub_totaal, Bedrijfsnaam, KVK, Rekeningummer, Btw_nummer, month_name, fac_number, date_today = process_excel_file(uploaded_file,client_info)
        df0 = pd.read_excel(uploaded_file, header=16)
        df0['Medewerker'][0] = df0['Medewerker'][0].lower()
        # Define the text information
        text_info = [
            ["Allround Care"],
            ["T.a.v. Administratie"],
            ["Wiersedreef 22"],
            ["3433 ZX Nieuwegein"],
            ["" , "", "", "", "","", "","", "","","",Bedrijfsnaam],
            ["" , "", "", "", "","", "","","","","", client_info.iloc[:,5][list(client_info.iloc[:,1]).index(df0['Medewerker'][0])]],
            ["" , "", "", "", "","", "","","","","", client_info.iloc[:,6][list(client_info.iloc[:,1]).index(df0['Medewerker'][0])]],
            [""], 
            ["" , "", "", "", "","", "","","","","", "KVK: " + str(KVK)],  
            ["" , "", "", "", "","", "","","","","", "Rekeningummer: " + str(Rekeningummer)],  
            ["" , "", "", "", "","", "","","","","", "Btw nummer: " + str(Btw_nummer)],  
            [""],
            [""],
            [""],
            [""],
            ["Maand:" + month_name, "" ,"", "" , "Factuurnummer:" + fac_number, "" ,"","", "Factuurdatum:" + date_today] 
            ]

        # Define the data for the row
        data = [
            [''],
            [''],
            ['','','','','','','','','Uurtarief maal uren', Totale_uren, Uurtarief, '€' + str(round(Uren_maal_tarief,2)) ],
            ['','','','','','','','','Totaal kilometervergoeding', Totale_km, Km_vergoeding, KM_maal_tarief],
            ['','','','','','','','','Sub-totaal', '', '', '€' + str(round(sub_totaal,2))],
            ['','','','','','','','','BTW%', '21%', '', '€' + str(round(0.21 * sub_totaal,2))],
            ['','','','','','','','','FACTUUR BEDRAG', '', '', '€' + str(round(1.21 * sub_totaal,2))],
            ['','','','',],
            ['','','','','','','','','Rekeningnummer', Rekeningummer, '', ''],
            [""],
            [""],
            ['Wij verzoeken u vriendelijk het bovenstaande bedrag binnen 30 dagen over te maken op bovenstaand rekeningnummer onder vermelding van het factuurnummer.', '', '', ''],
        ]
         # Create the dataframe
        df2 = pd.DataFrame(data)
        #df2.loc[-2] = ['Wij verzoeken u vriendelijk het bovenstaande bedrag binnen 30 dagen over te maken op bovenstaand rekeningnummer onder vermelding van het factuurnummer.', '', '', '']
        
        for i in range(len(df)):
            #st.write(df.iloc[:,2][i])
            if isinstance(df.iloc[:,2][i],str) == False:
             df.iloc[:,2][i] = round(float(df.iloc[:,2][i]),3)
            if isinstance(df.iloc[:,3][i],str) == False:    
             df.iloc[:,3][i] = round(float(df.iloc[:,3][i]),3)
            if isinstance(df.iloc[:,4][i],str) == False:
             df.iloc[:,4][i] = round(float(df.iloc[:,4][i]),3)
            if isinstance(df.iloc[:,5][i],str) == False:
             df.iloc[:,5][i] = round(float(df.iloc[:,5][i]),3)
            if isinstance(df.iloc[:,6][i],str) == False:
             df.iloc[:,6][i] = round(float(df.iloc[:,6][i]),3)
            if isinstance(df.iloc[:,7][i],str) == False:
             df.iloc[:,7][i] = round(float(df.iloc[:,7][i]),3)
            if isinstance(df.iloc[:,10][i],str) == False:    
             df.iloc[:,10][i] = round(float(df.iloc[:,10][i]),3)
        # Show the DataFrame with the calculated total amounts
        st.write('preview factuur')
        st.dataframe(df)

        st.write('Factuurbedrag totaal:', f'€{1.21 * sub_totaal:.2f}')
    
    #if st.button('Download de factuur'):
        # Create an ExcelWriter object
        excel_file_path = 'Factuur.xlsx'
        
        writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')

        # Write the text information to the sheet
        worksheet = writer.book.create_sheet('Sheet1')
        for row_idx, row_data in enumerate(text_info, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx).value = cell_value
        
        # Write the dataframes to the Excel file
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=len(text_info) + 1)
        df2.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=len(text_info) + 2 + len(df))
        
        # Close the Excel file
        writer.save()

        workbook = openpyxl.load_workbook('Factuur.xlsx')

        # select the sheet to modify
        sheet = workbook['Sheet1']
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['K'].width = 20
        sheet.column_dimensions['L'].width = 20
        # save the changes
        workbook.save('Factuur.xlsx')
        
        with open(excel_file_path, "rb") as file:
          btn=st.download_button(
          label="click me to download invoice",
          data=file,
          file_name=excel_file_path,
          mime="application/octet-stream"
          )
        #st.success(f'Download voltooid!')
        #df90 = pd.read_excel(excel_file_path)#input
        #df90.to_html("file.html")#to html
        xlsx2html('Factuur.xlsx', 'file.html')
        pdfkit.from_file("file.html", "Factuur.pdf")
        

        with open("Factuur.pdf", "rb") as file:
           btn=st.download_button(
           label="click me to download pdf",
           data=file,
           file_name="Factuur.pdf",
           mime="application/octet-stream"
           )
        if st.button('Verzend mail'):
         receiver = mail
         body = """Geachte, 

                In de bijlage treft u de factuur. 

                Voor vragen en/of opmerkingen kunt u contact met mij opnemen.


                Met vriendelijke groet,

                AAA"""
         filename = ["Factuur.pdf","Factuur.xlsx"]
         try:   
           yag = yagmail.SMTP(sender_email,sender_password)
           yag.send(
                 to=receiver,
                 subject='Factuur '+ month_name + ' ' + name ,
                 contents=body, 
                 attachments=filename,
                )
           st.success(f'E-mail succesvol verzonden!')
         except Exception as e:
             st.write(e)
if __name__ == '__main__':
    main()
